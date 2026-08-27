"""
One-off standalone script (Item #23 spec, §2.4): adds an Excel data-validation
dropdown to docs/Intake Template.xlsx's "If Enhancement -- Existing Service Name"
cell (Overview!C13), listing the real current services/<n>/ folder names.

Defense-in-depth alongside (not instead of) Ingestion Agent's/implementation_
coordinator.py's Layer 2 raise-on-mismatch backstop -- reduces typo risk at the
point of BA entry. This list is static and needs manual maintenance as new
services ship; the Layer 2 backstop is what catches the case where it drifts
out of date.

Usage: python scripts/add_existing_service_dropdown.py
"""

from __future__ import annotations

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

_TEMPLATE_PATH = "docs/Intake Template.xlsx"
_TARGET_CELL = "C13"
_CURRENT_SERVICES = ["REQ-2026-01", "REQ-2026-02", "REQ-2026-03"]


def main() -> None:
    wb = openpyxl.load_workbook(_TEMPLATE_PATH)
    ws = wb["Overview"]

    label = ws["B13"].value or ""
    if "existing service name" not in label.lower():
        raise ValueError(
            f"Overview!B13 does not look like the 'Existing Service Name' label "
            f"(found: {label!r}) -- template layout may have changed. Refusing "
            "to add a dropdown to the wrong cell."
        )

    formula = f'"{",".join(_CURRENT_SERVICES)}"'
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showDropDown=False,  # openpyxl quirk: False actually SHOWS the in-cell dropdown arrow
        showErrorMessage=False,  # advisory only -- do not hard-block a value not yet in this static list
        showInputMessage=True,
        promptTitle="Existing Service Name",
        prompt=(
            "Pick the exact services/<name>/ folder this Enhancement targets. "
            "If it's not listed (a newer service), type the exact folder name "
            "manually -- this list needs manual upkeep as new services ship."
        ),
    )
    dv.add(ws[_TARGET_CELL])
    ws.add_data_validation(dv)

    wb.save(_TEMPLATE_PATH)
    print(
        f"Added dropdown validation to Overview!{_TARGET_CELL} with options: "
        f"{_CURRENT_SERVICES}"
    )


if __name__ == "__main__":
    main()
