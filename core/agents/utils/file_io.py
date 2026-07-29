"""
FORGE file I/O helpers — read/write the artifact files that flow between pipeline stages.

Supported formats:
    - Excel (.xlsx) — reads the FORGE Intake Template structure
    - Markdown (.md) — plain read/write
    - YAML (.yaml/.yml) — read/write via PyYAML

Intake Template structure (as of Phase 1 design):
    Sheet "Instructions" — ignored (how-to guide for BAs, not parsed)
    Sheet "Overview"     — six named sections; each section header is in column A
                           (prefixed with a letter, e.g. "A — Request Identification").
                           Within each section, column B holds field labels and column C
                           holds BA input values. Title rows sit above the first section
                           — row numbers are not assumed.
    Sheet "Requirements" — tabular with a header row located by searching column A for
                           "Req #". The header is not assumed to be at row 1 — title rows
                           sit above it. Every row below the header where the
                           User Story / Requirement cell is non-empty is treated as a
                           real requirement. No example-row detection is applied: BAs are
                           instructed to replace or delete examples before submitting, so
                           any row present in a submitted file is real data.
                           Columns: Req #, Type, Priority, User Story / Requirement,
                           Acceptance Criteria, Notes / Constraints.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import openpyxl
import yaml

logger = logging.getLogger(__name__)

# ── Excel / Intake Template ────────────────────────────────────────────────────

# Column name mapping from the Requirements sheet header row (left-to-right order).
_REQ_COLUMNS = [
    "req_number",
    "type",
    "priority",
    "user_story",
    "acceptance_criteria",
    "notes",
]

# Index of the "User Story / Requirement" column within _REQ_COLUMNS.
# Rows where this cell is empty are skipped — they contain no actionable requirement.
_USER_STORY_COL_IDX = 3

# Maps lowercase substrings found in Overview column-A section headers to canonical
# snake_case keys. Substring match is used (case-insensitive) so prefixed labels like
# "A — Request Identification" are handled correctly without hardcoding exact cell text.
# The canonical keys are the guaranteed shape of the "overview" dict returned by read_xlsx().
_OVERVIEW_SECTION_KEYWORDS: dict[str, str] = {
    "request identification": "request_identification",
    "request type":           "request_type",
    "problem":                "problem_purpose",
    "success criteria":       "success_criteria_scope",
    "constraints":            "constraints_considerations",
    "additional context":     "additional_context",
}


def read_xlsx(path: str | Path) -> dict[str, Any]:
    """
    Parse a completed FORGE Intake Template spreadsheet.

    Args:
        path: Absolute or relative path to the .xlsx file.

    Returns:
        A dict with two keys:

        "overview" — dict with exactly these six canonical keys, each mapping to a
            dict of {field_label: field_value} pairs (column B = label, column C = value).
            An empty dict {} means the BA left that section blank.
                request_identification   — who is making the request, project/system name
                request_type             — new feature / enhancement / bug fix / etc.
                problem_purpose          — problem being solved and why it matters
                success_criteria_scope   — what done looks like and what is out of scope
                constraints_considerations — tech, time, resource, or policy constraints
                additional_context       — anything else the BA wants the team to know
            Example shape:
                {
                  "request_identification": {
                    "Request ID": "FORGE-2026-001",
                    "Request Title": "Client Portal — Document Upload",
                    ...
                  },
                  "request_type": {"Request Type": "Greenfield"},
                  ...
                }

        "requirements" — list of dicts, one per requirement row, with keys:
                req_number, type, priority, user_story, acceptance_criteria, notes
            Rows where user_story is empty are excluded; no example-row heuristic is
            applied (BAs are instructed to replace/delete examples before submitting).

    Raises:
        FileNotFoundError: If the file does not exist.
        KeyError: If the expected sheet names are not found.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Intake spreadsheet not found: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_names = {s.lower(): s for s in wb.sheetnames}

    # ── Overview sheet ─────────────────────────────────────────────────────────
    overview_key = next(
        (sheet_names[k] for k in sheet_names if "overview" in k),
        None,
    )
    if not overview_key:
        raise KeyError(f"No 'Overview' sheet found in {path.name}. Sheets: {wb.sheetnames}")

    ws_overview = wb[overview_key]
    overview: dict[str, dict] = {}
    current_section: str | None = None

    for row in ws_overview.iter_rows(values_only=True):
        cell_a = str(row[0] or "").strip()
        # Detect section headers by substring match (case-insensitive) so that
        # prefixed labels like "A — Request Identification" are handled correctly.
        canonical_key = next(
            (v for k, v in _OVERVIEW_SECTION_KEYWORDS.items() if k in cell_a.lower()),
            None,
        )
        if canonical_key:
            current_section = canonical_key
            overview[current_section] = {}
        elif current_section is not None:
            # Column B (index 1) = field label, column C (index 2) = field value.
            # Only add the pair when the label cell is non-empty.
            label = str(row[1] or "").strip() if len(row) > 1 else ""
            value = str(row[2] or "").strip() if len(row) > 2 else ""
            if label:
                overview[current_section][label] = value or None

    logger.debug("Parsed Overview sheet: %s sections found", len(overview))

    # ── Requirements sheet ─────────────────────────────────────────────────────
    req_key = next(
        (sheet_names[k] for k in sheet_names if "requirement" in k),
        None,
    )
    if not req_key:
        raise KeyError(f"No 'Requirements' sheet found in {path.name}. Sheets: {wb.sheetnames}")

    ws_req = wb[req_key]
    rows = list(ws_req.iter_rows(values_only=True))

    # Find the header row by searching column A (index 0) for "Req #".
    # The header is not assumed to be at row 1 — title rows sit above it in the template.
    header_row_idx = None
    for i, row in enumerate(rows):
        if row and "req #" in str(row[0] or "").strip().lower():
            header_row_idx = i
            break

    if header_row_idx is None:
        raise KeyError(
            f"Could not locate header row in the Requirements sheet of {path.name} "
            f"— expected 'Req #' in column A"
        )

    requirements: list[dict[str, Any]] = []
    for row in rows[header_row_idx + 1 :]:
        # Include the row only if the User Story / Requirement cell is non-empty.
        # No example-row heuristic: BAs are instructed to replace or delete the
        # pre-populated examples before submitting, so every present row is real data.
        user_story_val = row[_USER_STORY_COL_IDX] if len(row) > _USER_STORY_COL_IDX else None
        if user_story_val is None or str(user_story_val).strip() == "":
            continue

        req: dict[str, Any] = {}
        for idx, col_name in enumerate(_REQ_COLUMNS):
            req[col_name] = row[idx] if idx < len(row) else None
        requirements.append(req)

    logger.info(
        "Parsed Requirements sheet: %d requirement(s) found in %s",
        len(requirements),
        path.name,
    )

    return {"overview": overview, "requirements": requirements}


# ── Markdown ───────────────────────────────────────────────────────────────────

def read_markdown(path: str | Path) -> str:
    """
    Read a Markdown file and return its contents as a string.

    Args:
        path: Path to the .md file.

    Returns:
        File contents as a string.

    Raises:
        FileNotFoundError: If the file does not exist.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Markdown file not found: {path}")
    content = path.read_text(encoding="utf-8")
    logger.debug("Read markdown file: %s (%d chars)", path.name, len(content))
    return content


def write_markdown(path: str | Path, content: str) -> None:
    """
    Write a string to a Markdown file, creating parent directories as needed.

    Args:
        path: Destination path for the .md file.
        content: Markdown content to write.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")
    logger.info("Wrote markdown file: %s (%d chars)", path.name, len(content))


# ── YAML ───────────────────────────────────────────────────────────────────────

def read_yaml(path: str | Path) -> Any:
    """
    Parse a YAML file and return the resulting Python object.

    Args:
        path: Path to the .yaml or .yml file.

    Returns:
        Parsed Python object (typically a dict or list).

    Raises:
        FileNotFoundError: If the file does not exist.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"YAML file not found: {path}")
    with path.open(encoding="utf-8") as f:
        data = yaml.safe_load(f)
    logger.debug("Read YAML file: %s", path.name)
    return data


def write_yaml(path: str | Path, data: Any) -> None:
    """
    Serialize a Python object to a YAML file, creating parent directories as needed.

    Args:
        path: Destination path for the .yaml file.
        data: Python object to serialize (dict, list, etc.).
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        yaml.dump(data, f, default_flow_style=False, allow_unicode=True)
    logger.info("Wrote YAML file: %s", path.name)
